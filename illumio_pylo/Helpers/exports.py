from typing import Dict, Any, List, Optional, TypedDict, NotRequired, Union, Iterable

import xlsxwriter
import openpyxl
import csv
import os

import illumio_pylo as pylo


class ExcelHeader:
    def __init__(self, name: str, nice_name: Optional[str] = None, max_width: Optional[int] = None, wrap_text: Optional[bool] = None, is_url: [bool] = False, url_text: str = 'Link'):
        self.name = name
        self.nice_name:str = nice_name if nice_name is not None else name
        self.max_width = max_width
        self.wrap_text = wrap_text
        self.url_text = url_text
        self.is_url = is_url


class ExcelHeaderSet(list[ExcelHeader]):
    def __init__(self, headers: Optional[Iterable[str|ExcelHeader]]):
        super(ExcelHeaderSet, self).__init__()
        if headers is not None:
            for header in headers:
                self.append(header)

    def append(self, value: str|ExcelHeader):
        if type(value) is str:
            super(ExcelHeaderSet, self).append(ExcelHeader(name=value))
        elif type(value) is ExcelHeader:
            super(ExcelHeaderSet, self).append(value)
        else:
            raise pylo.PyloEx("ExcelHeaderSet.append() must be a string or a dict (ExcelHeader)")

    def _check_unique(self, new_header_name: str):
        if new_header_name in self:
            raise pylo.PyloEx("Header '{}' is already in the set".format(new_header_name))

    def extend(self, iterable):
        for item in iterable:
            self.append(item)

    def get_header(self, header_name: str) -> Optional[ExcelHeader]:
        for header in self:
            if header.name == header_name:
                return header
        return None


class ArrayToExport:

    def __init__(self, headers: List[str]):
        self._headers = headers
        self._columns_count = len(headers)
        self._lines = []

        self._headers_name_to_index = {}
        self._headers_index_to_name = []
        index = 0
        for header_name in headers:
            self._headers_name_to_index[header_name] = index
            self._headers_index_to_name.append(header_name)
            index += 1

    def columns_count(self):
        return len(self._headers)

    def lines_count(self):
        return len(self._lines)

    def add_line_from_object(self, record):
        new_line = []
        for header in self._headers:
            new_line.append(record.get(header))

        self._lines.append(new_line)

    def add_line_from_list_of_objects(self, list_of_objects):
        for record in list_of_objects:
            self.add_line_from_object(record)

    def add_line_from_list(self, line: list):
        if len(line) != self._columns_count:
            raise pylo.PyloEx("line length ({}) does not match the number of columns ({})".format(len(line), self._columns_count))
        self._lines.append(line)

    def write_to_csv(self, filename, delimiter=',', multivalues_cell_delimiter=' '):
        with open(filename, 'w', newline='') as csv_file:
            filewriter = csv.writer(csv_file, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_ALL)
            filewriter.writerow(self._headers)
            for line in self._lines:
                new_line = []
                for item in line:
                    if type(item) is list:
                        new_line.append(pylo.string_list_to_text(item, multivalues_cell_delimiter))
                    else:
                        new_line.append(item)
                filewriter.writerow(new_line)

    def write_to_excel(self, filename, worksheet_name='worksheet1', multivalues_cell_delimiter=' '):
        xls_workbook = xlsxwriter.Workbook(filename)
        cell_format = xls_workbook.add_format()
        cell_format.set_text_wrap()
        cell_format.set_valign('vcenter')
        xls_worksheet = xls_workbook.add_worksheet(worksheet_name)
        xls_headers = []
        xls_data = []
        header_index = 0
        for header in self._headers:
            xls_headers.append({'header': header, 'format': cell_format})
            header_index += 1

        # Building data array
        for line in self._lines:
            new_line = []
            for item in line:
                if type(item) is list:
                    new_line.append(pylo.string_list_to_text(item, multivalues_cell_delimiter))
                else:
                    new_line.append(item)
            xls_data.append(new_line)

        xls_table = xls_worksheet.add_table(0, 0, len(self._lines), len(self._headers)-1,
                                            {'header_row': True, 'data': xls_data, 'columns': xls_headers}
                                            )
        xls_worksheet.freeze_panes(1, 0)
        xls_workbook.close()


class ArraysToExcel:

    _sheets: Dict[str, 'ArraysToExcel.Sheet']

    class Sheet:
        def __init__(self, headers: ExcelHeaderSet, force_all_wrap_text=True, sheet_color: Optional[str] = None, order_by: Optional[List[str]] = None, multivalues_cell_delimiter=' '):
            self._headers: ExcelHeaderSet = headers
            self._columns_count = len(headers)
            self._lines = []
            self._columns_wrap = []
            self._color = sheet_color

            self._order_by = order_by

            self._multivalues_cell_delimiter = multivalues_cell_delimiter

            self._headers_name_to_index = {}
            self._headers_index_to_name = []
            index = 0

            for header in headers:
                self._columns_wrap.append(force_all_wrap_text)

                if header.nice_name is not None:
                    header.nice_name = header.name

                self._headers_index_to_name.append(header.name)
                self._headers_name_to_index[header.name] = index

                wrap = header.wrap_text
                if wrap is not None and not wrap:
                    self._columns_wrap[len(self._columns_wrap)-1] = False

                index += 1

        def write_to_csv(self, filename: str,
        ):
            headers: List[str] = []
            for header in self._headers:
                headers.append(header.name)

            headers_id: List[str] = []
            for header in self._headers:
                headers_id.append(header.name)


            exporter = ArrayToExport(headers)

            for line in self._lines:
                row = []
                for header in headers_id:
                    row.append(line[self._headers_name_to_index[header]])
                exporter.add_line_from_list(row)

            exporter.write_to_csv(filename)

        def columns_count(self):
            return len(self._headers)

        def lines_count(self):
            return len(self._lines)

        def reorder_lines(self, order_by: List[str]):
            self._lines = sorted(self._lines, key=lambda x: [x[self._headers_name_to_index[header_name]] for header_name in order_by])

        def add_line_from_object(self, record):
            new_line = []
            for header in self._headers:
                new_line.append(record.get(header.name))

            self._lines.append(new_line)

        def add_line_from_list_of_objects(self, list_of_objects):
            for record in list_of_objects:
                self.add_line_from_object(record)

        def add_line_from_list(self, line: list):
            if len(line) != self._columns_count:
                raise pylo.PyloEx("line length ({}) does not match the number of columns ({})".format(len(line), self._columns_count))
            self._lines.append(line)

        def add_to_document(self, xls_workbook: xlsxwriter.Workbook, sheet_name: str):

            def find_length(some_text: str) -> int:
                if type(some_text) is bool or some_text is None:
                    return 0
                if type(some_text) is int:
                    return len(str(some_text))

                str_length = 0
                split = some_text.split("\n")
                for part in split:
                    if len(part) > str_length:
                        str_length = len(part)

                return str_length


            # Data may need to be sorted
            if self._order_by is not None and len(self._order_by) > 0:
                self._lines = sorted(self._lines, key=lambda x: [x[self._headers_name_to_index[header_name]] for header_name in self._order_by])
                # print("********* Sorted by {}".format(self._order_by))

            xls_worksheet = xls_workbook.add_worksheet(sheet_name)
            if self._color is not None:
                xls_worksheet.tab_color = self._color
            xls_headers = []
            xls_data = []

            columns_max_width = []
            for header in self._headers:
                columns_max_width.append(0)



            # for each line, find the max length of string for each column and also add it to the data array
            for line in self._lines:
                new_line = []

                for item_index, item in enumerate(line):
                    if type(item) is list:
                        new_line.append(pylo.string_list_to_text(item, self._multivalues_cell_delimiter))
                    else:
                        if self._headers[item_index].is_url:
                            new_line.append('=HYPERLINK("{}", "{}")'.format(item,self._headers[item_index].url_text))
                        else:
                            new_line.append(item)
                            length = find_length(new_line[item_index])
                            if length > columns_max_width[item_index]:
                                columns_max_width[item_index] = length


                xls_data.append(new_line)

            header_index = 0
            for header in self._headers:
                cell_format = xls_workbook.add_format()
                cell_format.set_text_wrap(self._columns_wrap[header_index])
                cell_format.set_valign('vcenter')

                header_max_width_setting = None

                column_name = header.nice_name
                xls_headers.append({'header': column_name, 'format': cell_format})
                column_name_length = len(column_name) + 2 # add 2 for dropdown menus
                header_max_width_setting = header.max_width

                # default is to use width=longest string
                column_width = columns_max_width[header_index]+1

                if column_width < column_name_length:
                    column_width = column_name_length

                if header_max_width_setting is not None:
                    if header_max_width_setting is None or header_max_width_setting == 'auto':
                        pass
                    elif columns_max_width[header_index] > header_max_width_setting:
                        column_width = header_max_width_setting

                #print("column '{}' width={} vs setting={} vs calculated={}".format(header_index, column_width, header_max_width_setting, columns_max_width[header_index]))

                xls_worksheet.set_column(header_index, header_index, width=column_width*1.1)

                header_index += 1

            if len(self._lines) > 0:
                xls_table = xls_worksheet.add_table(0, 0, len(self._lines), len(self._headers)-1,
                                                {'header_row': True, 'data': xls_data, 'columns': xls_headers}
                                                )
            else:
                fake_data = []
                for header in self._headers:
                    fake_data.append(None)

                xls_table = xls_worksheet.add_table(0, 0, 1, len(self._headers)-1,
                                                    {'header_row': True, 'data': [fake_data], 'columns': xls_headers}
                                                    )
            xls_worksheet.freeze_panes(1, 0)

    def __init__(self):
        self._sheets = {}

    def create_sheet(self, name: str, headers: ExcelHeaderSet, force_all_wrap_text: bool = True, sheet_color: Optional[str] = None,
                     order_by: Optional[List[str]] = None, multivalues_cell_delimiter: str = ' ') -> Sheet:
        if name in self._sheets:
            pylo.PyloEx("A sheet named '{}' already exists".format(name))

        self._sheets[name] = ArraysToExcel.Sheet(headers, force_all_wrap_text=force_all_wrap_text,
                                                 sheet_color=sheet_color, order_by=order_by,
                                                 multivalues_cell_delimiter=multivalues_cell_delimiter)
        return self._sheets[name]

    def write_to_excel(self, filename, multivalues_cell_delimiter=' '):
        xls_workbook = xlsxwriter.Workbook(filename)

        for sheet_name, sheet_object in self._sheets.items():
            sheet_object.add_to_document(xls_workbook, sheet_name)

        xls_workbook.close()

    def add_line_from_object(self, record, sheet_name: str):
        self._sheets[sheet_name].add_line_from_object(record)

    def add_line_from_list_of_objects(self, list_of_objects, sheet_name: str):
        self._sheets[sheet_name].add_line_from_list_of_objects(list_of_objects)

    def add_line_from_list(self, line: list, sheet_name: str):
        self._sheets[sheet_name].add_line_from_list(line)



class CsvExcelToObject:

    def __init__(self, filename: str, expected_headers=None, csv_delimiter=',', csv_quotechar='"', strict_headers=False, excel_sheet_name=None):

        self._detected_headers = []
        self._header_index_to_name = []
        self._raw_lines = []
        self._objects = []
        self._empty_lines_count = 0

        if strict_headers and expected_headers is None:
            pylo.PyloEx("CSV/Excel file cannot use strict_headers mode without specifying expected_headers")

        if not os.path.exists(filename):
            raise pylo.PyloEx("File '{}' does not exist".format(filename))

        optional_headers = []
        mandatory_headers_dict = {}
        for header_infos in expected_headers:
            value = header_infos.get('optional')
            if value is None or value is True:
                optional_headers.append(header_infos)
            else:
                mandatory_headers_dict[header_infos['name']] = header_infos

        file_base, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()

        if file_extension == '.csv':
            with open(filename) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=csv_delimiter, quotechar=csv_quotechar)
                row_count = 0
                for row in csv_reader:

                    # this is Headers
                    if row_count == 0:
                        for item in row:
                            if item is None or len(item) < 1:
                                raise pylo.PyloEx('CSV headers has blank fields, this is not supported')
                            self._detected_headers.append(item.lower())
                            if strict_headers and item.lower() not in mandatory_headers_dict:
                                raise pylo.PyloEx("CSV/Excel headers have an unexpected header named '{}'".format(item))

                        missing_headers = mandatory_headers_dict.copy()
                        for header_name in self._detected_headers:
                            self._header_index_to_name.append(header_name)
                            if header_name in missing_headers:
                                del missing_headers[header_name]
                        if len(missing_headers) > 0:
                            raise pylo.PyloEx('CSV is missing the following mandatory headers: {}'.format(pylo.string_list_to_text(missing_headers.keys())))

                    # this is DATA
                    else:
                        if len(self._detected_headers) != len(row):
                            raise pylo.PyloEx('CSV line #{} doesnt have the same fields ({}) count than the headers ({}), is it empty?'.format(row_count+1,
                                                                                                                            len(self._detected_headers),
                                                                                                                     len(row)))

                        self._raw_lines.append(row)
                        new_object = {'*line*': row_count+1}
                        self._objects.append(new_object)
                        row_index = 0
                        for item in row:
                            new_object[self._detected_headers[row_index]] = item
                            row_index += 1

                        # handling missing optional columns
                        for opt_header in optional_headers:
                            if opt_header['name'] not in new_object:
                                if 'default' in opt_header:
                                    new_object[opt_header['name']] = opt_header['default']
                                else:
                                    new_object[opt_header['name']] = ''

                    row_count += 1
        elif file_extension == '.xlsx':
            #workbook = openpyxl.load_workbook(filename, read_only=True)
            workbook = openpyxl.load_workbook(filename)
            # print("workbook has {} worksheets".format(len(workbook.worksheets)))
            if len(workbook.worksheets) < 1:
                raise pylo.PyloEx("Excel file has no Worksheet")

            source_worksheet = workbook.worksheets[0]

            if excel_sheet_name is not None:
                source_worksheet = workbook.get_sheet_by_name(excel_sheet_name)
                if source_worksheet is None:
                    raise pylo.PyloEx("Cannot find a Worksheet named '{}' in Excel file '{}'".format(excel_sheet_name, filename))

            # print("\n\nmax_col {} max_row {}\n\n".format(source_worksheet.max_column,source_worksheet.max_row))

            for row_count in range(1, source_worksheet.max_row+1):
                # this is Headers
                if row_count == 1:
                    for col_index in range(1, source_worksheet.max_column+1):
                        item = source_worksheet.cell(row_count, col_index).value
                        if item is None or len(item) < 1:
                            raise pylo.PyloEx('Excel headers has blank fields, this is not supported')
                        self._detected_headers.append(item.lower())
                        if strict_headers and item.lower() not in mandatory_headers_dict:
                            raise pylo.PyloEx("CSV/Excel headers have an unexpected header named '{}'".format(item))

                    missing_headers = mandatory_headers_dict.copy()
                    for header_name in self._detected_headers:
                        self._header_index_to_name.append(header_name)
                        if header_name in missing_headers:
                            del missing_headers[header_name]
                    if len(missing_headers) > 0:
                        raise pylo.PyloEx('Excel file is missing the following mandatory headers: {}'.format(pylo.string_list_to_text(missing_headers.keys())))

                # This is DATA
                else:
                    new_object = {'*line*': row_count}

                    some_fields_have_data = False
                    for col_index in range(1, source_worksheet.max_column+1):
                        item = source_worksheet.cell(row_count, col_index).value
                        new_object[self._detected_headers[col_index-1]] = item
                        if item is not None and (type(item) is int or type(item) is bool or len(item) > 0):
                            some_fields_have_data = True

                    if not some_fields_have_data:
                        self._empty_lines_count += 1
                        continue

                    self._objects.append(new_object)

                    # handling missing optional columns
                    for opt_header in optional_headers:
                        if opt_header['name'] not in new_object:
                            if 'default' in opt_header:
                                new_object[opt_header['name']] = opt_header['default']
                            else:
                                new_object[opt_header['name']] = ''

        else:
            raise pylo.PyloEx("Unsupported file extension '{}' in filename {}".format(file_extension, filename))


    def count_lines(self):
        return len(self._objects)

    def count_empty_lines(self):
        return self._empty_lines_count

    def count_columns(self):
        return len(self._detected_headers)

    def objects(self) -> List[Dict[str, Union[str, int, bool, None]]]:
        return list(self._objects)

    def headers(self):
        return list(self._detected_headers)

    def save_to_csv(self, filename: str, fields_filter: List[Any]):
        headers = []
        for field in fields_filter:
            headers.append(field['name'])

        exporter = ArrayToExport(headers)

        for obj in self._objects:
            row = []
            for header in headers:
                row.append(obj.get(header))
            exporter.add_line_from_list(row)

        exporter.write_to_csv(filename)

    def save_to_excel(self, filename: str, fields_filter: list):
        headers = []
        for field in fields_filter:
            headers.append(field['name'])

        exporter = ArrayToExport(headers)

        for obj in self._objects:
            row = []
            for header in headers:
                row.append(obj.get(header))
            exporter.add_line_from_list(row)

        exporter.write_to_excel(filename)

